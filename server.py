from fastapi import FastAPI, APIRouter, HTTPException, Depends, status
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from pymongo import MongoClient
from pymongo.cursor import Cursor
# Compatibility shim: pymongo Cursor doesn't have .to_list() like motor does
Cursor.to_list = lambda self, n: list(self.limit(n))
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr
from typing import List, Optional
import uuid
from datetime import datetime, timedelta
from passlib.context import CryptContext
from jose import JWTError, jwt
import base64
import secrets

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection - lazy connect (Passenger fork-safe)
mongo_url = os.environ['MONGO_URL']
client = MongoClient(mongo_url, connect=False, serverSelectionTimeoutMS=15000)
db = client[os.environ.get('DB_NAME', 'networth_db')]

# JWT Configuration
SECRET_KEY = os.environ.get('SECRET_KEY')
if not SECRET_KEY:
    raise ValueError("SECRET_KEY environment variable is required")
ALGORITHM = "HS256"
ACCESS_TOKEN_EXPIRE_MINUTES = 60 * 24 * 7  # 7 days

# Password hashing
pwd_context = CryptContext(schemes=["bcrypt"], deprecated="auto")
security = HTTPBearer()

# Create the main app
app = FastAPI(title="Peers by NetWorth API")
api_router = APIRouter(prefix="/api")

# Admin emails - these users will automatically be admins
ADMIN_EMAILS = ["flaviusblaga@gmail.com"]

# ==================== MODELS ====================

class UserBase(BaseModel):
    email: EmailStr
    name: str
    bio: Optional[str] = ""
    headline: Optional[str] = ""
    location: Optional[str] = ""
    skills: List[str] = []
    experience: List[dict] = []
    language: str = "en"  # "en" or "ro"
    avatar: Optional[str] = None  # base64 image

class UserCreate(UserBase):
    password: str
    invite_code: Optional[str] = None

class UserLogin(BaseModel):
    email: EmailStr
    password: str

class UserUpdate(BaseModel):
    name: Optional[str] = None
    bio: Optional[str] = None
    headline: Optional[str] = None
    location: Optional[str] = None
    skills: Optional[List[str]] = None
    experience: Optional[List[dict]] = None
    language: Optional[str] = None
    avatar: Optional[str] = None  # base64 image

class UserResponse(UserBase):
    id: str
    created_at: datetime
    connections_count: int = 0
    avatar: Optional[str] = None
    is_admin: bool = False
    is_blocked: bool = False

class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: UserResponse

class PostCreate(BaseModel):
    content: str
    image: Optional[str] = None  # base64
    link: Optional[str] = None

class PostUpdate(BaseModel):
    content: Optional[str] = None
    link: Optional[str] = None

class CommentUpdate(BaseModel):
    content: str

class PostResponse(BaseModel):
    id: str
    user_id: str
    user_name: str
    user_headline: Optional[str] = ""
    user_avatar: Optional[str] = None
    content: str
    image: Optional[str] = None
    link: Optional[str] = None
    likes: List[str] = []
    comments: List[dict] = []
    created_at: datetime

class CommentCreate(BaseModel):
    content: str

class ConnectionRequest(BaseModel):
    to_user_id: str

class ConnectionResponse(BaseModel):
    id: str
    from_user_id: str
    from_user_name: str
    from_user_headline: Optional[str] = ""
    from_user_avatar: Optional[str] = None
    to_user_id: str
    to_user_name: str
    to_user_headline: Optional[str] = ""
    to_user_avatar: Optional[str] = None
    status: str  # pending, accepted, rejected
    created_at: datetime

class MessageCreate(BaseModel):
    to_user_id: str
    content: str

class MessageResponse(BaseModel):
    id: str
    from_user_id: str
    from_user_name: str
    to_user_id: str
    to_user_name: str
    content: str
    read: bool = False
    created_at: datetime

class ConversationResponse(BaseModel):
    user_id: str
    user_name: str
    user_headline: Optional[str] = ""
    user_avatar: Optional[str] = None
    last_message: str
    last_message_time: datetime
    unread_count: int = 0

# ==================== GROUP MODELS ====================

class GroupCreate(BaseModel):
    name: str
    member_ids: List[str] = []
    avatar: Optional[str] = None

class GroupUpdate(BaseModel):
    name: Optional[str] = None
    avatar: Optional[str] = None

class GroupMemberInfo(BaseModel):
    id: str
    name: str
    avatar: Optional[str] = None
    headline: Optional[str] = ""

class GroupResponse(BaseModel):
    id: str
    name: str
    avatar: Optional[str] = None
    owner_id: str
    owner_name: str
    members: List[GroupMemberInfo] = []
    member_count: int = 0
    last_message: Optional[str] = None
    last_message_time: Optional[datetime] = None
    unread_count: int = 0
    created_at: datetime

class GroupMessageCreate(BaseModel):
    content: str

class GroupMessageResponse(BaseModel):
    id: str
    group_id: str
    from_user_id: str
    from_user_name: str
    from_user_avatar: Optional[str] = None
    content: str
    created_at: datetime

class GroupMembersAdd(BaseModel):
    user_ids: List[str]

# ==================== ADMIN MODELS ====================

class ReportCreate(BaseModel):
    reported_user_id: Optional[str] = None
    reported_post_id: Optional[str] = None
    reason: str
    description: Optional[str] = ""

class ReportResponse(BaseModel):
    id: str
    reporter_id: str
    reporter_name: str
    reported_user_id: Optional[str] = None
    reported_user_name: Optional[str] = None
    reported_post_id: Optional[str] = None
    reason: str
    description: str
    status: str  # pending, resolved, dismissed
    created_at: datetime

class AdminStats(BaseModel):
    total_users: int
    total_posts: int
    total_connections: int
    total_messages: int
    pending_reports: int
    blocked_users: int
    new_users_today: int
    new_posts_today: int

class AdminUserUpdate(BaseModel):
    is_admin: Optional[bool] = None
    is_blocked: Optional[bool] = None

# ==================== HELPERS ====================

def verify_password(plain_password: str, hashed_password: str) -> bool:
    return pwd_context.verify(plain_password, hashed_password)

def get_password_hash(password: str) -> str:
    return pwd_context.hash(password)

def create_access_token(data: dict) -> str:
    to_encode = data.copy()
    expire = datetime.utcnow() + timedelta(minutes=ACCESS_TOKEN_EXPIRE_MINUTES)
    to_encode.update({"exp": expire})
    return jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)

def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        user_id: str = payload.get("sub")
        if user_id is None:
            raise HTTPException(status_code=401, detail="Invalid token")
    except JWTError:
        raise HTTPException(status_code=401, detail="Invalid token")
    
    user = db.users.find_one({"id": user_id})
    if user is None:
        raise HTTPException(status_code=401, detail="User not found")
    if user.get("is_blocked", False):
        raise HTTPException(status_code=403, detail="Account is blocked")
    return user

def get_admin_user(current_user: dict = Depends(get_current_user)):
    """Verify user is an admin"""
    is_admin = current_user.get("is_admin", False) or current_user.get("email") in ADMIN_EMAILS
    if not is_admin:
        raise HTTPException(status_code=403, detail="Admin access required")
    return current_user

def get_connections_count(user_id: str) -> int:
    count = db.connections.count_documents({
        "$or": [
            {"from_user_id": user_id, "status": "accepted"},
            {"to_user_id": user_id, "status": "accepted"}
        ]
    })
    return count

# ==================== AUTH ROUTES ====================

@api_router.post("/auth/register", response_model=TokenResponse)
def register(user_data: UserCreate):
    # Check if email exists
    existing = db.users.find_one({"email": user_data.email})
    if existing:
        raise HTTPException(status_code=400, detail="Email already registered")

    # Invite code required (admins exempt)
    used_code = None
    used_source = ""
    if user_data.email not in ADMIN_EMAILS:
        code = (user_data.invite_code or "").strip().upper()
        if not code:
            raise HTTPException(status_code=403, detail="Invite code required")
        inv = db.invite_codes.find_one({"code": code, "active": True})
        if not inv:
            raise HTTPException(status_code=403, detail="Invalid invite code")
        if inv.get("max_uses") and inv.get("used_count", 0) >= inv["max_uses"]:
            raise HTTPException(status_code=403, detail="Invite code already used up")
        db.invite_codes.update_one(
            {"code": code},
            {"$inc": {"used_count": 1}, "$push": {"used_by": user_data.email}}
        )
        used_code = code
        used_source = inv.get("note", "")
    else:
        used_source = "admin"
    
    # Create user
    user_id = str(uuid.uuid4())
    is_admin = user_data.email in ADMIN_EMAILS
    user_dict = {
        "id": user_id,
        "email": user_data.email,
        "name": user_data.name,
        "password_hash": get_password_hash(user_data.password),
        "bio": user_data.bio or "",
        "headline": user_data.headline or "",
        "location": user_data.location or "",
        "skills": user_data.skills or [],
        "experience": user_data.experience or [],
        "language": user_data.language or "en",
        "is_admin": is_admin,
        "is_blocked": False,
        "invite_code": used_code,
        "invite_source": used_source,
        "created_at": datetime.utcnow()
    }

    db.users.insert_one(user_dict)
    
    # Create token
    access_token = create_access_token({"sub": user_id})
    
    return TokenResponse(
        access_token=access_token,
        user=UserResponse(
            id=user_id,
            email=user_data.email,
            name=user_data.name,
            bio=user_dict["bio"],
            headline=user_dict["headline"],
            location=user_dict["location"],
            skills=user_dict["skills"],
            experience=user_dict["experience"],
            language=user_dict["language"],
            created_at=user_dict["created_at"],
            connections_count=0,
            is_admin=is_admin,
            is_blocked=False
        )
    )

@api_router.post("/auth/login", response_model=TokenResponse)
def login(credentials: UserLogin):
    user = db.users.find_one({"email": credentials.email})
    if not user or not verify_password(credentials.password, user["password_hash"]):
        raise HTTPException(status_code=401, detail="Invalid email or password")
    
    access_token = create_access_token({"sub": user["id"]})
    connections_count = get_connections_count(user["id"])
    is_admin = user.get("is_admin", False) or user.get("email") in ADMIN_EMAILS
    
    return TokenResponse(
        access_token=access_token,
        user=UserResponse(
            id=user["id"],
            email=user["email"],
            name=user["name"],
            bio=user.get("bio", ""),
            headline=user.get("headline", ""),
            location=user.get("location", ""),
            skills=user.get("skills", []),
            experience=user.get("experience", []),
            language=user.get("language", "en"),
            created_at=user["created_at"],
            connections_count=connections_count,
            avatar=user.get("avatar"),
            is_admin=is_admin,
            is_blocked=user.get("is_blocked", False)
        )
    )

@api_router.get("/auth/me", response_model=UserResponse)
def get_me(current_user: dict = Depends(get_current_user)):
    connections_count = get_connections_count(current_user["id"])
    is_admin = current_user.get("is_admin", False) or current_user.get("email") in ADMIN_EMAILS
    return UserResponse(
        id=current_user["id"],
        email=current_user["email"],
        name=current_user["name"],
        bio=current_user.get("bio", ""),
        headline=current_user.get("headline", ""),
        location=current_user.get("location", ""),
        skills=current_user.get("skills", []),
        experience=current_user.get("experience", []),
        language=current_user.get("language", "en"),
        created_at=current_user["created_at"],
        connections_count=connections_count,
        avatar=current_user.get("avatar"),
        is_admin=is_admin,
        is_blocked=current_user.get("is_blocked", False)
    )

@api_router.put("/auth/me", response_model=UserResponse)
def update_me(update_data: UserUpdate, current_user: dict = Depends(get_current_user)):
    update_dict = {k: v for k, v in update_data.dict().items() if v is not None}
    if update_dict:
        db.users.update_one({"id": current_user["id"]}, {"$set": update_dict})
    
    updated_user = db.users.find_one({"id": current_user["id"]})
    connections_count = get_connections_count(current_user["id"])
    is_admin = updated_user.get("is_admin", False) or updated_user.get("email") in ADMIN_EMAILS
    
    return UserResponse(
        id=updated_user["id"],
        email=updated_user["email"],
        name=updated_user["name"],
        bio=updated_user.get("bio", ""),
        headline=updated_user.get("headline", ""),
        location=updated_user.get("location", ""),
        skills=updated_user.get("skills", []),
        experience=updated_user.get("experience", []),
        language=updated_user.get("language", "en"),
        created_at=updated_user["created_at"],
        connections_count=connections_count,
        avatar=updated_user.get("avatar"),
        is_admin=is_admin,
        is_blocked=updated_user.get("is_blocked", False)
    )

# ==================== USER ROUTES ====================

@api_router.get("/users", response_model=List[UserResponse])
def get_users(search: Optional[str] = None, current_user: dict = Depends(get_current_user)):
    query = {"id": {"$ne": current_user["id"]}, "is_blocked": {"$ne": True}}
    if search:
        query["$or"] = [
            {"name": {"$regex": search, "$options": "i"}},
            {"headline": {"$regex": search, "$options": "i"}},
            {"skills": {"$elemMatch": {"$regex": search, "$options": "i"}}}
        ]
    
    users = list(db.users.find(query).limit(100))
    result = []
    for user in users:
        connections_count = get_connections_count(user["id"])
        is_admin = user.get("is_admin", False) or user.get("email") in ADMIN_EMAILS
        result.append(UserResponse(
            id=user["id"],
            email=user["email"],
            name=user["name"],
            bio=user.get("bio", ""),
            headline=user.get("headline", ""),
            location=user.get("location", ""),
            skills=user.get("skills", []),
            experience=user.get("experience", []),
            language=user.get("language", "en"),
            created_at=user["created_at"],
            connections_count=connections_count,
            avatar=user.get("avatar")
        ))
    return result

@api_router.get("/users/{user_id}", response_model=UserResponse)
def get_user(user_id: str, current_user: dict = Depends(get_current_user)):
    user = db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    connections_count = get_connections_count(user["id"])
    return UserResponse(
        id=user["id"],
        email=user["email"],
        name=user["name"],
        bio=user.get("bio", ""),
        headline=user.get("headline", ""),
        location=user.get("location", ""),
        skills=user.get("skills", []),
        experience=user.get("experience", []),
        language=user.get("language", "en"),
        created_at=user["created_at"],
        connections_count=connections_count,
        avatar=user.get("avatar")
    )

# ==================== POST ROUTES ====================

@api_router.post("/posts", response_model=PostResponse)
def create_post(post_data: PostCreate, current_user: dict = Depends(get_current_user)):
    post_id = str(uuid.uuid4())
    post_dict = {
        "id": post_id,
        "user_id": current_user["id"],
        "user_name": current_user["name"],
        "user_headline": current_user.get("headline", ""),
        "content": post_data.content,
        "image": post_data.image,
        "link": post_data.link,
        "likes": [],
        "comments": [],
        "created_at": datetime.utcnow()
    }
    db.posts.insert_one(post_dict)
    return _enrich_post(post_dict)


def _enrich_post(post: dict) -> PostResponse:
    """Attach current user avatar (and comment authors' avatars) fetched from users collection."""
    # Post author avatar
    user_avatar = None
    author = db.users.find_one({"id": post.get("user_id")}, {"avatar": 1, "headline": 1, "name": 1})
    if author:
        user_avatar = author.get("avatar")
        # Also refresh headline & name (may have changed)
        post["user_headline"] = author.get("headline", post.get("user_headline", ""))
        post["user_name"] = author.get("name", post.get("user_name", ""))
    post["user_avatar"] = user_avatar

    # Enrich comments with authors' avatars
    comments = post.get("comments", []) or []
    if comments:
        author_ids = list({c.get("user_id") for c in comments if c.get("user_id")})
        if author_ids:
            authors_docs = list(db.users.find(
                {"id": {"$in": author_ids}},
                {"id": 1, "avatar": 1, "name": 1, "_id": 0}
            ))
            author_map = {a["id"]: a for a in authors_docs}
            for c in comments:
                a = author_map.get(c.get("user_id")) or {}
                c["user_avatar"] = a.get("avatar")
                if a.get("name"):
                    c["user_name"] = a["name"]
    return PostResponse(**post)


@api_router.get("/posts", response_model=List[PostResponse])
def get_posts(current_user: dict = Depends(get_current_user)):
    # Get user's connections
    connections = db.connections.find({
        "$or": [
            {"from_user_id": current_user["id"], "status": "accepted"},
            {"to_user_id": current_user["id"], "status": "accepted"}
        ]
    }).to_list(1000)
    
    connection_ids = set()
    for conn in connections:
        if conn["from_user_id"] == current_user["id"]:
            connection_ids.add(conn["to_user_id"])
        else:
            connection_ids.add(conn["from_user_id"])
    
    # Include own posts and connections' posts
    connection_ids.add(current_user["id"])
    
    posts = list(db.posts.find({"user_id": {"$in": list(connection_ids)}}).sort("created_at", -1).limit(100))
    return [_enrich_post(post) for post in posts]

@api_router.get("/posts/all", response_model=List[PostResponse])
def get_all_posts(current_user: dict = Depends(get_current_user)):
    posts = list(db.posts.find().sort("created_at", -1).limit(100))
    return [_enrich_post(post) for post in posts]

@api_router.get("/posts/user/{user_id}", response_model=List[PostResponse])
def get_user_posts(user_id: str, current_user: dict = Depends(get_current_user)):
    posts = list(db.posts.find({"user_id": user_id}).sort("created_at", -1).limit(100))
    return [_enrich_post(post) for post in posts]

@api_router.post("/posts/{post_id}/like")
def like_post(post_id: str, current_user: dict = Depends(get_current_user)):
    post = db.posts.find_one({"id": post_id})
    if not post:
        raise HTTPException(status_code=404, detail="Post not found")
    
    if current_user["id"] in post.get("likes", []):
        # Unlike
        db.posts.update_one({"id": post_id}, {"$pull": {"likes": current_user["id"]}})
        return {"message": "Unliked", "liked": False}
    else:
        # Like
        db.posts.update_one({"id": post_id}, {"$push": {"likes": current_user["id"]}})
        return {"message": "Liked", "liked": True}

@api_router.post("/posts/{post_id}/comment", response_model=PostResponse)
def add_comment(post_id: str, comment_data: CommentCreate, current_user: dict = Depends(get_current_user)):
    post = db.posts.find_one({"id": post_id})
    if not post:
        raise HTTPException(status_code=404, detail="Post not found")
    
    comment = {
        "id": str(uuid.uuid4()),
        "user_id": current_user["id"],
        "user_name": current_user["name"],
        "content": comment_data.content,
        "created_at": datetime.utcnow().isoformat()
    }
    
    db.posts.update_one({"id": post_id}, {"$push": {"comments": comment}})
    updated_post = db.posts.find_one({"id": post_id})
    return _enrich_post(updated_post)

@api_router.put("/posts/{post_id}/comments/{comment_id}", response_model=PostResponse)
def update_comment(post_id: str, comment_id: str, comment_data: CommentUpdate, current_user: dict = Depends(get_current_user)):
    post = db.posts.find_one({"id": post_id})
    if not post:
        raise HTTPException(status_code=404, detail="Post not found")

    comment = next((c for c in post.get("comments", []) if c.get("id") == comment_id), None)
    if not comment:
        raise HTTPException(status_code=404, detail="Comment not found")
    if comment.get("user_id") != current_user["id"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    content = comment_data.content.strip()
    if not content:
        raise HTTPException(status_code=400, detail="Comment content is required")

    db.posts.update_one(
        {"id": post_id, "comments.id": comment_id},
        {"$set": {"comments.$.content": content}}
    )

    updated_post = db.posts.find_one({"id": post_id})
    return _enrich_post(updated_post)

@api_router.put("/posts/{post_id}", response_model=PostResponse)
def update_post(post_id: str, post_data: PostUpdate, current_user: dict = Depends(get_current_user)):
    post = db.posts.find_one({"id": post_id})
    if not post:
        raise HTTPException(status_code=404, detail="Post not found")
    if post["user_id"] != current_user["id"]:
        raise HTTPException(status_code=403, detail="Not authorized")

    update_dict = {}
    if post_data.content is not None:
        update_dict["content"] = post_data.content
    if post_data.link is not None:
        update_dict["link"] = post_data.link
    if update_dict:
        db.posts.update_one({"id": post_id}, {"$set": update_dict})

    updated_post = db.posts.find_one({"id": post_id})
    return _enrich_post(updated_post)

@api_router.delete("/posts/{post_id}")
def delete_post(post_id: str, current_user: dict = Depends(get_current_user)):
    post = db.posts.find_one({"id": post_id})
    if not post:
        raise HTTPException(status_code=404, detail="Post not found")
    if post["user_id"] != current_user["id"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    db.posts.delete_one({"id": post_id})
    return {"message": "Post deleted"}

# ==================== CONNECTION ROUTES ====================

# ==================== CONNECTION ROUTES ====================

def _enrich_connection_avatars(connections: list) -> list:
    """Attach from_user_avatar and to_user_avatar to connection dicts."""
    if not connections:
        return []
    user_ids = set()
    for c in connections:
        user_ids.add(c.get("from_user_id"))
        user_ids.add(c.get("to_user_id"))
    user_ids.discard(None)
    users_cursor = db.users.find(
        {"id": {"$in": list(user_ids)}},
        {"id": 1, "avatar": 1, "_id": 0}
    )
    avatar_map = {u["id"]: u.get("avatar") for u in users_cursor}
    for c in connections:
        c["from_user_avatar"] = avatar_map.get(c.get("from_user_id"))
        c["to_user_avatar"] = avatar_map.get(c.get("to_user_id"))
    return connections


@api_router.post("/connections", response_model=ConnectionResponse)
def create_connection_request(request: ConnectionRequest, current_user: dict = Depends(get_current_user)):
    # Check if connection already exists
    existing = db.connections.find_one({
        "$or": [
            {"from_user_id": current_user["id"], "to_user_id": request.to_user_id},
            {"from_user_id": request.to_user_id, "to_user_id": current_user["id"]}
        ]
    })
    if existing:
        raise HTTPException(status_code=400, detail="Connection request already exists")
    
    to_user = db.users.find_one({"id": request.to_user_id})
    if not to_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    conn_id = str(uuid.uuid4())
    conn_dict = {
        "id": conn_id,
        "from_user_id": current_user["id"],
        "from_user_name": current_user["name"],
        "from_user_headline": current_user.get("headline", ""),
        "to_user_id": request.to_user_id,
        "to_user_name": to_user["name"],
        "to_user_headline": to_user.get("headline", ""),
        "status": "pending",
        "created_at": datetime.utcnow()
    }
    
    db.connections.insert_one(conn_dict)
    conn_dict["from_user_avatar"] = current_user.get("avatar")
    conn_dict["to_user_avatar"] = to_user.get("avatar")
    return ConnectionResponse(**conn_dict)

@api_router.get("/connections", response_model=List[ConnectionResponse])
def get_connections(current_user: dict = Depends(get_current_user)):
    connections = db.connections.find({
        "$or": [
            {"from_user_id": current_user["id"], "status": "accepted"},
            {"to_user_id": current_user["id"], "status": "accepted"}
        ]
    }).to_list(1000)
    connections = _enrich_connection_avatars(connections)
    return [ConnectionResponse(**conn) for conn in connections]

@api_router.get("/connections/pending", response_model=List[ConnectionResponse])
def get_pending_connections(current_user: dict = Depends(get_current_user)):
    connections = db.connections.find({
        "to_user_id": current_user["id"],
        "status": "pending"
    }).to_list(100)
    connections = _enrich_connection_avatars(connections)
    return [ConnectionResponse(**conn) for conn in connections]

@api_router.get("/connections/sent", response_model=List[ConnectionResponse])
def get_sent_connections(current_user: dict = Depends(get_current_user)):
    connections = db.connections.find({
        "from_user_id": current_user["id"],
        "status": "pending"
    }).to_list(100)
    connections = _enrich_connection_avatars(connections)
    return [ConnectionResponse(**conn) for conn in connections]

@api_router.put("/connections/{connection_id}/accept")
def accept_connection(connection_id: str, current_user: dict = Depends(get_current_user)):
    connection = db.connections.find_one({"id": connection_id})
    if not connection:
        raise HTTPException(status_code=404, detail="Connection not found")
    if connection["to_user_id"] != current_user["id"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    db.connections.update_one({"id": connection_id}, {"$set": {"status": "accepted"}})
    return {"message": "Connection accepted"}

@api_router.put("/connections/{connection_id}/reject")
def reject_connection(connection_id: str, current_user: dict = Depends(get_current_user)):
    connection = db.connections.find_one({"id": connection_id})
    if not connection:
        raise HTTPException(status_code=404, detail="Connection not found")
    if connection["to_user_id"] != current_user["id"]:
        raise HTTPException(status_code=403, detail="Not authorized")
    
    db.connections.delete_one({"id": connection_id})
    return {"message": "Connection rejected"}

@api_router.delete("/connections/{user_id}")
def remove_connection(user_id: str, current_user: dict = Depends(get_current_user)):
    result = db.connections.delete_one({
        "$or": [
            {"from_user_id": current_user["id"], "to_user_id": user_id, "status": "accepted"},
            {"from_user_id": user_id, "to_user_id": current_user["id"], "status": "accepted"}
        ]
    })
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Connection not found")
    return {"message": "Connection removed"}

@api_router.get("/connections/status/{user_id}")
def get_connection_status(user_id: str, current_user: dict = Depends(get_current_user)):
    connection = db.connections.find_one({
        "$or": [
            {"from_user_id": current_user["id"], "to_user_id": user_id},
            {"from_user_id": user_id, "to_user_id": current_user["id"]}
        ]
    })
    
    if not connection:
        return {"status": "none", "connection_id": None}
    
    is_sender = connection["from_user_id"] == current_user["id"]
    return {
        "status": connection["status"],
        "connection_id": connection["id"],
        "is_sender": is_sender
    }

# ==================== MESSAGE ROUTES ====================

@api_router.post("/messages", response_model=MessageResponse)
def send_message(message_data: MessageCreate, current_user: dict = Depends(get_current_user)):
    to_user = db.users.find_one({"id": message_data.to_user_id})
    if not to_user:
        raise HTTPException(status_code=404, detail="User not found")
    
    msg_id = str(uuid.uuid4())
    msg_dict = {
        "id": msg_id,
        "from_user_id": current_user["id"],
        "from_user_name": current_user["name"],
        "to_user_id": message_data.to_user_id,
        "to_user_name": to_user["name"],
        "content": message_data.content,
        "read": False,
        "created_at": datetime.utcnow()
    }
    
    db.messages.insert_one(msg_dict)
    return MessageResponse(**msg_dict)

@api_router.get("/messages/conversations", response_model=List[ConversationResponse])
def get_conversations(current_user: dict = Depends(get_current_user)):
    # Get all messages involving the user
    messages = list(db.messages.find({
        "$or": [
            {"from_user_id": current_user["id"]},
            {"to_user_id": current_user["id"]}
        ]
    }).sort("created_at", -1).limit(1000))
    
    conversations = {}
    for msg in messages:
        other_user_id = msg["to_user_id"] if msg["from_user_id"] == current_user["id"] else msg["from_user_id"]
        other_user_name = msg["to_user_name"] if msg["from_user_id"] == current_user["id"] else msg["from_user_name"]
        
        if other_user_id not in conversations:
            # Get user headline and avatar
            other_user = db.users.find_one({"id": other_user_id})
            headline = other_user.get("headline", "") if other_user else ""
            avatar = other_user.get("avatar") if other_user else None

            # Count unread
            unread = db.messages.count_documents({
                "from_user_id": other_user_id,
                "to_user_id": current_user["id"],
                "read": False
            })

            conversations[other_user_id] = ConversationResponse(
                user_id=other_user_id,
                user_name=other_user_name,
                user_headline=headline,
                user_avatar=avatar,
                last_message=msg["content"],
                last_message_time=msg["created_at"],
                unread_count=unread
            )
    
    return list(conversations.values())

@api_router.get("/messages/{user_id}", response_model=List[MessageResponse])
def get_messages_with_user(user_id: str, current_user: dict = Depends(get_current_user)):
    messages = list(db.messages.find({
        "$or": [
            {"from_user_id": current_user["id"], "to_user_id": user_id},
            {"from_user_id": user_id, "to_user_id": current_user["id"]}
        ]
    }).sort("created_at", 1).limit(1000))
    
    # Mark messages as read
    db.messages.update_many(
        {"from_user_id": user_id, "to_user_id": current_user["id"], "read": False},
        {"$set": {"read": True}}
    )
    
    return [MessageResponse(**msg) for msg in messages]

# ==================== GROUP CHAT ROUTES ====================

def _serialize_group(group: dict, current_user_id: str) -> GroupResponse:
    """Convert DB group document to GroupResponse."""
    member_ids = group.get("member_ids", [])
    members_docs = list(db.users.find(
        {"id": {"$in": member_ids}},
        {"id": 1, "name": 1, "avatar": 1, "headline": 1, "_id": 0}
    ))
    members = [
        GroupMemberInfo(
            id=m["id"],
            name=m.get("name", ""),
            avatar=m.get("avatar"),
            headline=m.get("headline", "")
        )
        for m in members_docs
    ]
    owner = db.users.find_one({"id": group["owner_id"]}, {"name": 1})
    owner_name = owner.get("name", "") if owner else ""

    last_msg_doc = db.group_messages.find_one(
        {"group_id": group["id"]},
        sort=[("created_at", -1)]
    )
    last_message = last_msg_doc["content"] if last_msg_doc else None
    last_message_time = last_msg_doc["created_at"] if last_msg_doc else None

    # Unread count: messages after current user's last-read timestamp
    reads = group.get("read_state", {}) or {}
    last_read = reads.get(current_user_id)
    unread_filter = {"group_id": group["id"], "from_user_id": {"$ne": current_user_id}}
    if last_read:
        unread_filter["created_at"] = {"$gt": last_read}
    unread_count = db.group_messages.count_documents(unread_filter)

    return GroupResponse(
        id=group["id"],
        name=group["name"],
        avatar=group.get("avatar"),
        owner_id=group["owner_id"],
        owner_name=owner_name,
        members=members,
        member_count=len(member_ids),
        last_message=last_message,
        last_message_time=last_message_time,
        unread_count=unread_count,
        created_at=group["created_at"],
    )


@api_router.post("/groups", response_model=GroupResponse)
def create_group(payload: GroupCreate, current_user: dict = Depends(get_current_user)):
    name = payload.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="Group name is required")
    if len(name) > 60:
        raise HTTPException(status_code=400, detail="Group name too long")

    # Sanitize member ids: must be connected to current user (accepted connection) or the user themselves
    member_ids = list({mid for mid in payload.member_ids if mid and mid != current_user["id"]})
    if member_ids:
        connected = db.connections.find({
            "$or": [
                {"from_user_id": current_user["id"], "to_user_id": {"$in": member_ids}, "status": "accepted"},
                {"to_user_id": current_user["id"], "from_user_id": {"$in": member_ids}, "status": "accepted"},
            ]
        })
        allowed = set()
        for conn in connected:
            allowed.add(conn["to_user_id"] if conn["from_user_id"] == current_user["id"] else conn["from_user_id"])
        member_ids = [m for m in member_ids if m in allowed]

    # Always include the owner
    all_members = list({current_user["id"], *member_ids})

    group_id = str(uuid.uuid4())
    now = datetime.utcnow()
    group_doc = {
        "id": group_id,
        "name": name,
        "avatar": payload.avatar,
        "owner_id": current_user["id"],
        "member_ids": all_members,
        "read_state": {current_user["id"]: now},
        "created_at": now,
    }
    db.groups.insert_one(group_doc)
    return _serialize_group(group_doc, current_user["id"])


@api_router.get("/groups", response_model=List[GroupResponse])
def list_my_groups(current_user: dict = Depends(get_current_user)):
    groups = list(db.groups.find({"member_ids": current_user["id"]}).sort("created_at", -1).limit(200))
    result = [_serialize_group(g, current_user["id"]) for g in groups]
    # Sort by last_message_time desc, then by created_at desc
    result.sort(key=lambda g: (g.last_message_time or g.created_at), reverse=True)
    return result


@api_router.get("/groups/{group_id}", response_model=GroupResponse)
def get_group(group_id: str, current_user: dict = Depends(get_current_user)):
    group = db.groups.find_one({"id": group_id})
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    if current_user["id"] not in group.get("member_ids", []):
        raise HTTPException(status_code=403, detail="You are not a member of this group")
    return _serialize_group(group, current_user["id"])


@api_router.put("/groups/{group_id}", response_model=GroupResponse)
def update_group(group_id: str, payload: GroupUpdate, current_user: dict = Depends(get_current_user)):
    group = db.groups.find_one({"id": group_id})
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    if group["owner_id"] != current_user["id"]:
        raise HTTPException(status_code=403, detail="Only the group owner can update it")
    updates = {}
    if payload.name is not None:
        name = payload.name.strip()
        if not name:
            raise HTTPException(status_code=400, detail="Group name cannot be empty")
        if len(name) > 60:
            raise HTTPException(status_code=400, detail="Group name too long")
        updates["name"] = name
    if payload.avatar is not None:
        updates["avatar"] = payload.avatar or None
    if updates:
        db.groups.update_one({"id": group_id}, {"$set": updates})
        group.update(updates)
    return _serialize_group(group, current_user["id"])


@api_router.delete("/groups/{group_id}")
def delete_group(group_id: str, current_user: dict = Depends(get_current_user)):
    group = db.groups.find_one({"id": group_id})
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    if group["owner_id"] != current_user["id"]:
        raise HTTPException(status_code=403, detail="Only the group owner can delete it")
    db.groups.delete_one({"id": group_id})
    db.group_messages.delete_many({"group_id": group_id})
    return {"success": True}


@api_router.post("/groups/{group_id}/members", response_model=GroupResponse)
def add_members(group_id: str, payload: GroupMembersAdd, current_user: dict = Depends(get_current_user)):
    group = db.groups.find_one({"id": group_id})
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    if current_user["id"] not in group.get("member_ids", []):
        raise HTTPException(status_code=403, detail="You are not a member of this group")

    new_ids = [uid for uid in payload.user_ids if uid and uid not in group.get("member_ids", [])]
    if new_ids:
        # Only allow users that current user is connected with
        connected = db.connections.find({
            "$or": [
                {"from_user_id": current_user["id"], "to_user_id": {"$in": new_ids}, "status": "accepted"},
                {"to_user_id": current_user["id"], "from_user_id": {"$in": new_ids}, "status": "accepted"},
            ]
        })
        allowed = set()
        for conn in connected:
            allowed.add(conn["to_user_id"] if conn["from_user_id"] == current_user["id"] else conn["from_user_id"])
        new_ids = [uid for uid in new_ids if uid in allowed]

    if new_ids:
        db.groups.update_one(
            {"id": group_id},
            {"$addToSet": {"member_ids": {"$each": new_ids}}}
        )
        group["member_ids"] = list({*group.get("member_ids", []), *new_ids})
    return _serialize_group(group, current_user["id"])


@api_router.delete("/groups/{group_id}/members/{user_id}", response_model=GroupResponse)
def remove_member(group_id: str, user_id: str, current_user: dict = Depends(get_current_user)):
    group = db.groups.find_one({"id": group_id})
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    if current_user["id"] not in group.get("member_ids", []):
        raise HTTPException(status_code=403, detail="You are not a member of this group")

    is_owner = group["owner_id"] == current_user["id"]
    is_self = user_id == current_user["id"]

    if not is_owner and not is_self:
        raise HTTPException(status_code=403, detail="Only the owner can remove other members")

    if is_owner and is_self:
        # Owner leaves: transfer ownership to next member, or delete if last
        remaining = [m for m in group.get("member_ids", []) if m != user_id]
        if remaining:
            db.groups.update_one(
                {"id": group_id},
                {"$set": {"owner_id": remaining[0]}, "$pull": {"member_ids": user_id}}
            )
        else:
            db.groups.delete_one({"id": group_id})
            db.group_messages.delete_many({"group_id": group_id})
            return {"id": group_id, "name": group["name"], "owner_id": user_id,
                    "owner_name": current_user.get("name", ""), "members": [],
                    "member_count": 0, "created_at": group["created_at"]}
    else:
        db.groups.update_one({"id": group_id}, {"$pull": {"member_ids": user_id}})

    group = db.groups.find_one({"id": group_id})
    return _serialize_group(group, current_user["id"])


@api_router.get("/groups/{group_id}/messages", response_model=List[GroupMessageResponse])
def get_group_messages(group_id: str, current_user: dict = Depends(get_current_user)):
    group = db.groups.find_one({"id": group_id})
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    if current_user["id"] not in group.get("member_ids", []):
        raise HTTPException(status_code=403, detail="You are not a member of this group")

    messages = list(db.group_messages.find({"group_id": group_id}).sort("created_at", 1).limit(1000))
    # Mark as read for this user
    db.groups.update_one(
        {"id": group_id},
        {"$set": {f"read_state.{current_user['id']}": datetime.utcnow()}}
    )
    return [GroupMessageResponse(**m) for m in messages]


@api_router.post("/groups/{group_id}/messages", response_model=GroupMessageResponse)
def send_group_message(group_id: str, payload: GroupMessageCreate, current_user: dict = Depends(get_current_user)):
    group = db.groups.find_one({"id": group_id})
    if not group:
        raise HTTPException(status_code=404, detail="Group not found")
    if current_user["id"] not in group.get("member_ids", []):
        raise HTTPException(status_code=403, detail="You are not a member of this group")

    content = payload.content.strip()
    if not content:
        raise HTTPException(status_code=400, detail="Message content is required")

    msg = {
        "id": str(uuid.uuid4()),
        "group_id": group_id,
        "from_user_id": current_user["id"],
        "from_user_name": current_user.get("name", ""),
        "from_user_avatar": current_user.get("avatar"),
        "content": content,
        "created_at": datetime.utcnow(),
    }
    db.group_messages.insert_one(msg)
    # Update sender's read timestamp
    db.groups.update_one(
        {"id": group_id},
        {"$set": {f"read_state.{current_user['id']}": msg["created_at"]}}
    )
    msg.pop("_id", None)
    return GroupMessageResponse(**msg)


# ==================== HEALTH CHECK ====================

@api_router.get("/health")
def health_check():
    return {"status": "ok", "timestamp": datetime.utcnow().isoformat()}

# ==================== ADMIN ROUTES ====================

@api_router.get("/admin/stats", response_model=AdminStats)
def get_admin_stats(admin_user: dict = Depends(get_admin_user)):
    today = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    
    total_users = db.users.count_documents({})
    total_posts = db.posts.count_documents({})
    total_connections = db.connections.count_documents({"status": "accepted"})
    total_messages = db.messages.count_documents({})
    pending_reports = db.reports.count_documents({"status": "pending"})
    blocked_users = db.users.count_documents({"is_blocked": True})
    new_users_today = db.users.count_documents({"created_at": {"$gte": today}})
    new_posts_today = db.posts.count_documents({"created_at": {"$gte": today}})
    
    return AdminStats(
        total_users=total_users,
        total_posts=total_posts,
        total_connections=total_connections,
        total_messages=total_messages,
        pending_reports=pending_reports,
        blocked_users=blocked_users,
        new_users_today=new_users_today,
        new_posts_today=new_posts_today
    )

@api_router.get("/admin/users", response_model=List[UserResponse])
def get_all_users_admin(admin_user: dict = Depends(get_admin_user)):
    users = list(db.users.find().sort("created_at", -1).limit(500))
    result = []
    for user in users:
        connections_count = get_connections_count(user["id"])
        is_admin = user.get("is_admin", False) or user.get("email") in ADMIN_EMAILS
        result.append(UserResponse(
            id=user["id"],
            email=user["email"],
            name=user["name"],
            bio=user.get("bio", ""),
            headline=user.get("headline", ""),
            location=user.get("location", ""),
            skills=user.get("skills", []),
            experience=user.get("experience", []),
            language=user.get("language", "en"),
            created_at=user["created_at"],
            connections_count=connections_count,
            avatar=user.get("avatar"),
            is_admin=is_admin,
            is_blocked=user.get("is_blocked", False)
        ))
    return result

@api_router.put("/admin/users/{user_id}")
def update_user_admin(user_id: str, update_data: AdminUserUpdate, admin_user: dict = Depends(get_admin_user)):
    user = db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    update_dict = {}
    if update_data.is_admin is not None:
        update_dict["is_admin"] = update_data.is_admin
    if update_data.is_blocked is not None:
        update_dict["is_blocked"] = update_data.is_blocked
    
    if update_dict:
        db.users.update_one({"id": user_id}, {"$set": update_dict})
    
    return {"message": "User updated successfully"}

@api_router.delete("/admin/users/{user_id}")
def delete_user_admin(user_id: str, admin_user: dict = Depends(get_admin_user)):
    user = db.users.find_one({"id": user_id})
    if not user:
        raise HTTPException(status_code=404, detail="User not found")
    
    # Delete user and all their data
    db.users.delete_one({"id": user_id})
    db.posts.delete_many({"user_id": user_id})
    db.connections.delete_many({"$or": [{"from_user_id": user_id}, {"to_user_id": user_id}]})
    db.messages.delete_many({"$or": [{"from_user_id": user_id}, {"to_user_id": user_id}]})
    
    return {"message": "User deleted successfully"}

class InviteCreate(BaseModel):
    max_uses: Optional[int] = None
    note: Optional[str] = ""

@api_router.post("/admin/invites")
def create_invite(payload: InviteCreate, admin_user: dict = Depends(get_admin_user)):
    alphabet = "ABCDEFGHJKLMNPQRSTUVWXYZ23456789"
    code = "".join(secrets.choice(alphabet) for _ in range(8))
    doc = {
        "code": code,
        "active": True,
        "max_uses": payload.max_uses,
        "used_count": 0,
        "used_by": [],
        "note": payload.note or "",
        "created_by": admin_user["email"],
        "created_at": datetime.utcnow()
    }
    db.invite_codes.insert_one(doc)
    doc.pop("_id", None)
    return doc

@api_router.get("/admin/invites")
def list_invites(admin_user: dict = Depends(get_admin_user)):
    return list(db.invite_codes.find({}, {"_id": 0}).sort("created_at", -1).limit(200))

@api_router.delete("/admin/invites/{code}")
def deactivate_invite(code: str, admin_user: dict = Depends(get_admin_user)):
    result = db.invite_codes.update_one({"code": code.upper()}, {"$set": {"active": False}})
    if result.matched_count == 0:
        raise HTTPException(status_code=404, detail="Invite code not found")
    return {"message": "Invite code deactivated"}

@api_router.get("/admin/members/export")
def export_members(admin_user: dict = Depends(get_admin_user)):
    from fastapi.responses import StreamingResponse
    from openpyxl import Workbook
    from openpyxl.styles import Font
    import io

    # Fallback map for members registered before invite tracking: email -> (code, note)
    code_map = {}
    for inv in db.invite_codes.find({}, {"_id": 0, "code": 1, "note": 1, "used_by": 1}):
        for em in inv.get("used_by", []):
            code_map.setdefault(em, (inv["code"], inv.get("note", "")))

    wb = Workbook()
    ws = wb.active
    ws.title = "Members"
    headers = ["Name", "Email", "Headline", "Location", "Language",
               "Registered at", "Invite code", "Source", "Admin", "Blocked"]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)

    for u in db.users.find().sort("created_at", 1):
        code = u.get("invite_code") or ""
        source = u.get("invite_source") or ""
        if not code and u.get("email") in code_map:
            code, mapped_note = code_map[u["email"]]
            source = source or mapped_note
        created = u.get("created_at")
        ws.append([
            u.get("name", ""),
            u.get("email", ""),
            u.get("headline", ""),
            u.get("location", ""),
            u.get("language", ""),
            created.strftime("%Y-%m-%d %H:%M") if created else "",
            code,
            source,
            "yes" if u.get("is_admin") else "",
            "yes" if u.get("is_blocked") else "",
        ])

    widths = [22, 30, 26, 18, 10, 17, 13, 22, 8, 8]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = w

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = "peers-members-" + datetime.utcnow().strftime("%Y-%m-%d") + ".xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@api_router.get("/admin/posts")
def get_all_posts_admin(admin_user: dict = Depends(get_admin_user)):
    posts = list(db.posts.find().sort("created_at", -1).limit(500))
    return posts

@api_router.delete("/admin/posts/{post_id}")
def delete_post_admin(post_id: str, admin_user: dict = Depends(get_admin_user)):
    post = db.posts.find_one({"id": post_id})
    if not post:
        raise HTTPException(status_code=404, detail="Post not found")
    
    db.posts.delete_one({"id": post_id})
    return {"message": "Post deleted successfully"}

@api_router.post("/reports", response_model=ReportResponse)
def create_report(report_data: ReportCreate, current_user: dict = Depends(get_current_user)):
    report_id = str(uuid.uuid4())
    
    reported_user_name = None
    if report_data.reported_user_id:
        reported_user = db.users.find_one({"id": report_data.reported_user_id})
        reported_user_name = reported_user["name"] if reported_user else None
    
    report_dict = {
        "id": report_id,
        "reporter_id": current_user["id"],
        "reporter_name": current_user["name"],
        "reported_user_id": report_data.reported_user_id,
        "reported_user_name": reported_user_name,
        "reported_post_id": report_data.reported_post_id,
        "reason": report_data.reason,
        "description": report_data.description or "",
        "status": "pending",
        "created_at": datetime.utcnow()
    }
    
    db.reports.insert_one(report_dict)
    return ReportResponse(**report_dict)

@api_router.get("/admin/reports", response_model=List[ReportResponse])
def get_reports_admin(status: Optional[str] = None, admin_user: dict = Depends(get_admin_user)):
    query = {}
    if status:
        query["status"] = status
    
    reports = list(db.reports.find(query).sort("created_at", -1).limit(200))
    return [ReportResponse(**report) for report in reports]

@api_router.put("/admin/reports/{report_id}")
def update_report_admin(report_id: str, status: str, admin_user: dict = Depends(get_admin_user)):
    report = db.reports.find_one({"id": report_id})
    if not report:
        raise HTTPException(status_code=404, detail="Report not found")
    
    if status not in ["pending", "resolved", "dismissed"]:
        raise HTTPException(status_code=400, detail="Invalid status")
    
    db.reports.update_one({"id": report_id}, {"$set": {"status": status}})
    return {"message": "Report updated successfully"}

# Include the router
app.include_router(api_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
def shutdown_db_client():
    client.close()
